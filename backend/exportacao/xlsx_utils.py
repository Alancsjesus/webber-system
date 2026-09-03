"""
Utilitários de exportação em Excel (.xlsx). Único ponto do projeto que usa
openpyxl — mantido separado de pdf_utils.py por natureza de biblioteca.
"""
from decimal import Decimal

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from django.http import HttpResponse


def resposta_xlsx(workbook: openpyxl.Workbook, filename: str) -> HttpResponse:
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


def gerar_xlsx_relatorio_itens_plano_aplicacao(itens: list, filename: str) -> HttpResponse:
    """
    Gera XLSX do Relatório de Itens do Plano de Aplicação FESP. `itens` é uma
    lista de ItemPlanoAplicacao já anotados com `.executado` (ver
    modulo_fesp.indicadores).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Itens'

    headers = [
        'Plano', 'Exercício', 'Eixo', 'Órgão Gestor', 'Órgão Beneficiário',
        'Meta', 'Natureza', 'Bem/Serviço', 'Status', 'Executado', 'Valor Total (R$)',
    ]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    for item in itens:
        plano = item.meta_especifica.plano
        ws.append([
            plano.numero,
            plano.exercicio_fiscal,
            plano.ementa,
            plano.org_id.sigla if plano.org_id else '',
            item.org_beneficiaria.sigla,
            item.meta_especifica.titulo,
            item.get_natureza_display(),
            item.bem_servico,
            item.get_status_display(),
            'Sim' if item.executado else 'Não',
            float(item.valor_total_estimado or Decimal('0')),
        ])

    for i, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(len(header) + 2, 14)

    return resposta_xlsx(wb, filename)


def gerar_xlsx_painel_tramitacao(itens_painel: list, filename: str) -> HttpResponse:
    """
    Gera XLSX do Painel Gerencial de Tramitação. `itens_painel` é a lista de
    dicts já resolvida por `modulo_tramitacao.estagio` + `ProcessoTramitacao`
    manual sem DFD — ver `PainelTramitacaoView._itens()`.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Tramitação'

    headers = ['Setor', 'Processo SEI', 'Objeto', 'Fonte(s) de Recurso', 'Fase', 'Data de Entrada na Fase']
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    for p in itens_painel:
        ws.append([
            p['setor'],
            p['numero_sei'],
            p['objeto'],
            ', '.join(p['fontes_recurso_nomes']),
            p['fase_atual'],
            p['data_entrada_fase'].strftime('%d/%m/%Y') if p['data_entrada_fase'] else '',
        ])

    for i, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(len(header) + 2, 16)

    return resposta_xlsx(wb, filename)
